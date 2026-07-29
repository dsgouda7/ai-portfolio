"""XLSX extractor — serializes each sheet to 'Header: value' prose rows."""

from __future__ import annotations

from pathlib import Path

from .base import BaseExtractor


class XlsxExtractor(BaseExtractor):
    def extract(self, path: Path) -> str:
        return self._safe(self._read, path)

    def _read(self, path: Path) -> str:
        import openpyxl  # type: ignore[import]

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        parts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # First non-empty row is treated as headers
            headers = [str(c) if c is not None else "" for c in rows[0]]
            parts.append(f"[Sheet: {sheet_name}]")

            for row in rows[1:]:
                cells = [str(c) if c is not None else "" for c in row]
                # Serialize as "Header=value; Header=value" — prose-friendly
                pairs = [f"{h}={v}" for h, v in zip(headers, cells) if h and v]
                if pairs:
                    parts.append("; ".join(pairs))

        wb.close()
        return "\n".join(parts)
