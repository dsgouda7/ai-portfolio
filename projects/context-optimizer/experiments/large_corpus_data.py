"""Utilities for preparing and loading large benchmark corpora.

This module intentionally keeps IO logic separate from experiment scoring logic so
dataset preparation can run independently (and in parallel).
"""

from __future__ import annotations

import random
import re
from pathlib import Path
from urllib.request import urlopen


DEFAULT_GUTENBERG_URLS = [
    "https://www.gutenberg.org/cache/epub/2600/pg2600.txt",   # War and Peace
    "https://www.gutenberg.org/cache/epub/1184/pg1184.txt",   # The Count of Monte Cristo
]


def _strip_gutenberg_boilerplate(text: str) -> str:
    start_patterns = [
        r"\*\*\*\s*START OF (?:THE|THIS) PROJECT GUTENBERG EBOOK.*?\*\*\*",
        r"START OF THE PROJECT GUTENBERG EBOOK",
    ]
    end_patterns = [
        r"\*\*\*\s*END OF (?:THE|THIS) PROJECT GUTENBERG EBOOK.*?\*\*\*",
        r"END OF THE PROJECT GUTENBERG EBOOK",
    ]

    stripped = text
    for pat in start_patterns:
        m = re.search(pat, stripped, flags=re.IGNORECASE | re.DOTALL)
        if m:
            stripped = stripped[m.end() :]
            break

    for pat in end_patterns:
        m = re.search(pat, stripped, flags=re.IGNORECASE | re.DOTALL)
        if m:
            stripped = stripped[: m.start()]
            break

    return stripped.strip()


def download_gutenberg_books(
    output_dir: Path,
    urls: list[str] | None = None,
    timeout_s: int = 45,
) -> Path:
    """Download one or more Gutenberg books and save a combined corpus file."""
    output_dir.mkdir(parents=True, exist_ok=True)
    urls = urls or list(DEFAULT_GUTENBERG_URLS)

    print(f"[Gutenberg] Downloading {len(urls)} books...")
    combined_parts: list[str] = []
    for idx, url in enumerate(urls, start=1):
        print(f"[Gutenberg] Downloading book {idx}/{len(urls)}: {url}")
        with urlopen(url, timeout=timeout_s) as resp:
            raw = resp.read().decode("utf-8", errors="replace")
        clean = _strip_gutenberg_boilerplate(raw)
        combined_parts.append(f"\n\n### BOOK_{idx} SOURCE: {url}\n\n{clean}")

    combined = "\n".join(combined_parts)
    out_file = output_dir / "combined_gutenberg.txt"
    out_file.write_text(combined, encoding="utf-8")
    print(f"[Gutenberg] ✓ Saved {out_file} ({out_file.stat().st_size / 1024 / 1024:.1f} MB)")
    return out_file


def build_gutenberg_corpus_lines(
    corpus_path: Path,
    max_segments: int = 14000,
) -> list[str]:
    """Convert a large Gutenberg text corpus into segment-addressable lines."""
    text = corpus_path.read_text(encoding="utf-8", errors="replace")
    raw_blocks = [b.strip() for b in re.split(r"\n\s*\n", text) if b.strip()]

    # Gutenberg texts often contain short heading blocks. Aggregate them into
    # bounded segments so the corpus remains large and retrieval-relevant.
    lines: list[str] = []
    segment_parts: list[str] = []
    segment_chars = 0
    seg_idx = 1

    for block in raw_blocks:
        compact = " ".join(block.split())
        if not compact:
            continue
        segment_parts.append(compact)
        segment_chars += len(compact)

        if segment_chars >= 900:
            joined = " ".join(segment_parts)
            lines.append(f"g:seg{seg_idx:05d}: {joined}")
            seg_idx += 1
            segment_parts = []
            segment_chars = 0
            if len(lines) >= max_segments:
                break

    if segment_parts and len(lines) < max_segments:
        joined = " ".join(segment_parts)
        lines.append(f"g:seg{seg_idx:05d}: {joined}")

    # Fallback: if parsing still yields too few segments, split by sentence-like chunks.
    if len(lines) < 20:
        sentences = [s.strip() for s in re.split(r"(?<=[.!?])\s+", text) if s.strip()]
        lines = []
        chunk: list[str] = []
        chunk_chars = 0
        seg_idx = 1
        for sentence in sentences:
            chunk.append(sentence)
            chunk_chars += len(sentence)
            if chunk_chars >= 900:
                lines.append(f"g:seg{seg_idx:05d}: {' '.join(chunk)}")
                seg_idx += 1
                chunk = []
                chunk_chars = 0
                if len(lines) >= max_segments:
                    break
        if chunk and len(lines) < max_segments:
            lines.append(f"g:seg{seg_idx:05d}: {' '.join(chunk)}")

    return lines


def generate_large_excel_mock(
    output_path: Path,
    target_mb: int = 120,
    seed: int = 42,
) -> Path:
    """Generate a large synthetic XLSX file suitable for retrieval/analysis benchmarks."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for large Excel generation. Install with `pip install openpyxl`."
        ) from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    rng = random.Random(seed)

    rows_per_mb = 7000
    total_rows_target = max(120_000, target_mb * rows_per_mb)

    print(f"[Excel] Generating ~{target_mb}MB Excel file with ~{total_rows_target:,} rows...")

    headers = [
        "event_ts",
        "customer_id",
        "order_id",
        "region",
        "channel",
        "status",
        "error_code",
        "amount",
        "tax",
        "cost",
        "margin",
        "latency_ms",
        "score",
        "risk",
        "campaign",
        "device",
        "country",
        "city",
        "agent_name",
        "ticket_summary",
    ]

    wb = Workbook(write_only=True)

    max_sheet_rows = 1_000_000
    rows_written = 0
    sheet_idx = 1

    while rows_written < total_rows_target:
        ws = wb.create_sheet(title=f"sheet_{sheet_idx}")
        ws.append(headers)
        sheet_idx += 1

        rows_this_sheet = min(max_sheet_rows - 1, total_rows_target - rows_written)
        print(f"[Excel] Writing sheet {sheet_idx-1}: {rows_this_sheet:,} rows...")

        for _ in range(rows_this_sheet):
            day = rng.randint(1, 28)
            hour = rng.randint(0, 23)
            minute = rng.randint(0, 59)
            region = rng.choice(["NA", "EU", "APAC", "LATAM", "MEA"])
            channel = rng.choice(["web", "mobile", "store", "partner"])
            status = rng.choice(["ok", "warning", "failed", "retry"])
            err = rng.choice(["none", "E101", "E210", "E512", "E770", "E999"])
            amount = round(rng.uniform(5, 5000), 2)
            tax = round(amount * rng.uniform(0.02, 0.2), 2)
            cost = round(amount * rng.uniform(0.3, 0.95), 2)
            margin = round(amount - cost - tax, 2)
            latency = rng.randint(20, 4000)
            score = round(rng.uniform(0, 1), 4)
            risk = round(rng.uniform(0, 1), 4)
            campaign = rng.choice(["spring", "summer", "fall", "winter", "evergreen"])
            device = rng.choice(["ios", "android", "desktop", "tablet"])
            country = rng.choice(["US", "GB", "IN", "DE", "BR", "CA", "AU"])
            city = rng.choice(["nyc", "london", "mumbai", "berlin", "sao-paulo", "toronto", "sydney"])
            agent = rng.choice(["alice", "bob", "carol", "dave", "eve", "mallory"])
            ticket = (
                f"order anomaly for {region}/{channel}; status={status}; "
                f"latency={latency}ms; campaign={campaign}; device={device}"
            )

            ws.append(
                [
                    f"2026-06-{day:02d}T{hour:02d}:{minute:02d}:00Z",
                    rng.randint(100_000, 9_999_999),
                    rng.randint(1_000_000, 99_999_999),
                    region,
                    channel,
                    status,
                    err,
                    amount,
                    tax,
                    cost,
                    margin,
                    latency,
                    score,
                    risk,
                    campaign,
                    device,
                    country,
                    city,
                    agent,
                    ticket,
                ]
            )
            rows_written += 1

    print(f"[Excel] Saving workbook to {output_path}...")
    wb.save(output_path)
    print(f"[Excel] ✓ Generated {output_path} ({output_path.stat().st_size / 1024 / 1024:.1f} MB)")
    return output_path


def build_excel_corpus_lines(
    excel_path: Path,
    max_rows: int = 250_000,
) -> list[str]:
    """Read a large XLSX and emit retrieval-friendly text lines."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for reading XLSX corpus lines. Install with `pip install openpyxl`."
        ) from exc

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    lines: list[str] = []

    for ws in wb.worksheets:
        row_iter = ws.iter_rows(values_only=True)
        headers = next(row_iter, None)
        if not headers:
            continue
        cols = [str(c) for c in headers]

        for row_idx, row in enumerate(row_iter, start=2):
            if len(lines) >= max_rows:
                return lines
            row_map = {cols[i]: row[i] for i in range(min(len(cols), len(row)))}
            lines.append(
                "sheet={sheet} row={row} region={region} channel={channel} status={status} "
                "error={error} amount={amount} margin={margin} latency_ms={latency} risk={risk} summary={summary}".format(
                    sheet=ws.title,
                    row=row_idx,
                    region=row_map.get("region", "na"),
                    channel=row_map.get("channel", "na"),
                    status=row_map.get("status", "na"),
                    error=row_map.get("error_code", "na"),
                    amount=row_map.get("amount", "na"),
                    margin=row_map.get("margin", "na"),
                    latency=row_map.get("latency_ms", "na"),
                    risk=row_map.get("risk", "na"),
                    summary=row_map.get("ticket_summary", ""),
                )
            )

    return lines
